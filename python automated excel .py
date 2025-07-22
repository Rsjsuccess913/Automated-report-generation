#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

data = {
    'EmployeeID': [101, 102, 103, 104, 105, 106, 107],
    'Department': ['Sales', 'HR', 'Sales', 'IT', 'HR', 'IT', 'IT'],
    'JoiningDate': ['2019-01-10', '2020-03-15', '2018-07-25', '2019-11-10', '2021-01-05', '2019-12-05', '2020-06-20'],
    'Salary': [70000, 65000, 72000, 80000, 63000, 77000, 75000],
    'Attrition': ['No', 'Yes', 'No', 'No', 'Yes', 'No', 'No']
}
df = pd.DataFrame(data)
df['JoiningDate'] = pd.to_datetime(df['JoiningDate'])

summary_df = df.groupby('Department').agg(
    AvgSalary=('Salary', 'mean'),
    AttritionRate=('Attrition', lambda x: (x == 'Yes').mean())
).reset_index()

detailed_df = df.copy()

output_file = 'HR_Report_Automated.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    detailed_df.to_excel(writer, index=False, sheet_name='Employee Details')
    summary_df.to_excel(writer, index=False, sheet_name='Summary')

wb = load_workbook(output_file)

ws = wb['Summary']

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

col_widths = {'A': 15, 'B': 12, 'C': 15}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

chart = BarChart()
chart.title = "Average Salary by Department"
chart.y_axis.title = 'Avg Salary'
chart.x_axis.title = 'Department'

data_ref = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

ws.add_chart(chart, "E2")

wb.save(output_file)
print(f"Report generated: {output_file}")


# In[ ]:




